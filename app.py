# -*- coding: utf-8 -*-
"""
电子料台账协作管理平台
Leader创建任务 → 分配人员 → 人员填写4大模块 → 提交完成 → Leader审批
运行: pip install flask && python app.py
"""

import os, sys, sqlite3, io, json, secrets, tempfile
# Fix Windows GBK encoding for print
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except:
        pass
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, redirect,
    url_for, session, flash, g, jsonify, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# App Config
# ============================================================
app = Flask(__name__)
app.secret_key = 'material-ledger-secret-key-2024'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
DATABASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database.db')
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ============================================================
# Field Definitions (from 电子料台账表.xlsx, 39 fields / 4 modules)
# ============================================================
MODULES = [
    {
        'key': 'component_info', 'name': '料件信息', 'icon': '📋',
        'fields': [
            {'col': 'unique_id',        'label': '唯一编号',             'type': 'text', 'required': True},
            {'col': 'product',          'label': '产品（量产+潜在）',     'type': 'text', 'required': True},
            {'col': 'material_code',    'label': '物料编码',             'type': 'text', 'required': True},
            {'col': 'material_name',    'label': '物料名称',             'type': 'text', 'required': True},
            {'col': 'material_model',   'label': '物料型号',             'type': 'text', 'required': True},
            {'col': 'material_desc',    'label': '物料描述',             'type': 'textarea'},
        ]
    },
    {
        'key': 'bom_usage', 'name': 'BOM用量', 'icon': '📊',
        'fields': [
            {'col': 'bom_2k_v2',          'label': '2K V2',                    'type': 'number'},
            {'col': 'bom_2k_v5',          'label': '2K V5',                    'type': 'number'},
            {'col': 'bom_s21_b0_ti',      'label': 'S2.1 B0 TI版本',           'type': 'number'},
            {'col': 'bom_s20_b0_rtw',     'label': 'S2.0 B0 锐泰微版本',       'type': 'number'},
            {'col': 'bom_f_project',      'label': 'F项目（角用量）',           'type': 'number'},
            {'col': 'bom_forward',        'label': '前向用量',                  'type': 'number'},
            {'col': 'bom_general',        'label': 'Bom用量（通用）',           'type': 'number'},
            {'col': 'bom_24t24r_sat',     'label': '24T24R卫星方案',           'type': 'number'},
            {'col': 'bom_24t24r_soc',     'label': '24T24R SOC方案',           'type': 'number'},
            {'col': 'bom_sfr2_4d',        'label': 'SFR2-4D',                  'type': 'number'},
            {'col': 'bom_nio_8t8r',       'label': 'NIO 8T8R SFR2-4D-S2',     'type': 'number'},
            {'col': 'bom_horizon_8t8r',   'label': '地平线 8T8R SFR2-4D-S3',  'type': 'number'},
            {'col': 'bom_total',          'label': '总计',                      'type': 'number', 'readonly': True},
        ]
    },
    {
        'key': 'commercial', 'name': '商流信息', 'icon': '💰',
        'fields': [
            {'col': 'unit',                    'label': '单位',                'type': 'text', 'required': True},
            {'col': 'manufacturer',            'label': '品牌/制造商',          'type': 'text', 'required': True},
            {'col': 'supplier',                'label': '供应商',              'type': 'text', 'required': True},
            {'col': 'currency',                'label': '币别',                'type': 'select', 'options': ['RMB','USD','EUR','JPY','TWD']},
            {'col': 'unit_price_untaxed',      'label': '单价未税',            'type': 'text'},
            {'col': 'unit_price_taxed',        'label': '单价含税',            'type': 'text'},
            {'col': 'spq',                     'label': 'SPQ（标准包装）',      'type': 'text'},
            {'col': 'moq',                     'label': 'MOQ（最小起订）',      'type': 'text'},
            {'col': 'lead_time',               'label': 'L/T（交期）',         'type': 'text'},
            {'col': 'production_time',         'label': '量产时间',            'type': 'text'},
            {'col': 'factory_monthly_output',  'label': '原厂月出货量',         'type': 'text'},
            {'col': 'distributor_monthly_output','label': '代理商月出货量',      'type': 'text'},
            {'col': 'automotive_applications',  'label': '车厂应用实绩',         'type': 'textarea'},
            {'col': 'wafer_origin',            'label': '晶圆产地',            'type': 'text'},
            {'col': 'packaging_test_origin',   'label': '封测产地',            'type': 'text'},
            {'col': 'shipping_origin',         'label': '出货地',              'type': 'text'},
            {'col': 'peak_monthly_capacity',   'label': '峰值月产能',           'type': 'text'},
            {'col': 'cancel_window_days',      'label': 'Cancel Window Days', 'type': 'text'},
        ]
    },
    {
        'key': 'other', 'name': '其他', 'icon': '📝',
        'fields': [
            {'col': 'update_frequency',    'label': '更新频次',     'type': 'select', 'options': ['日','周','月度','季度','半年','年度']},
            {'col': 'attachment_note',     'label': '附件（报价依据）','type': 'textarea'},
        ]
    },
]

STATUS_MAP = {
    'assigned':    {'label': '待填写', 'color': '#2a78d6', 'bg': '#e8f0fe'},
    'in_progress': {'label': '填写中', 'color': '#eda100', 'bg': '#fff8e6'},
    'submitted':   {'label': '已提交', 'color': '#eb6834', 'bg': '#fff0ec'},
    'completed':   {'label': '已完成', 'color': '#0ca30c', 'bg': '#ecfdf3'},
}

# Leader 在创建任务时填写模块1（料件信息），成员填写模块2-4
LEADER_MODULES = [MODULES[0]]  # 料件信息
MEMBER_MODULES = [MODULES[1], MODULES[2], MODULES[3]]  # BOM用量、商流信息、其他

# ============================================================
# Database
# ============================================================
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(e):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def init_db():
    """Create tables and seed users."""
    db = sqlite3.connect(DATABASE)
    db.execute("PRAGMA foreign_keys = ON")

    # Build task columns
    task_cols = []
    for mod in MODULES:
        for f in mod['fields']:
            task_cols.append(f"{f['col']} TEXT DEFAULT ''")

    db.executescript(f"""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        display_name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'member' CHECK(role IN ('leader','member')),
        password_changed INTEGER DEFAULT 0,
        failed_attempts INTEGER DEFAULT 0,
        locked_until TEXT DEFAULT '',
        permissions TEXT DEFAULT '{{}}',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        {','.join(task_cols)},
        assigned_to INTEGER REFERENCES users(id),
        created_by INTEGER REFERENCES users(id),
        status TEXT NOT NULL DEFAULT 'assigned' CHECK(status IN ('assigned','in_progress','submitted','completed')),
        notes TEXT DEFAULT '',
        due_date TEXT DEFAULT '',
        attachment_file TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS activity_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
        user_id INTEGER REFERENCES users(id),
        action TEXT NOT NULL,
        comment TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status);
    CREATE INDEX IF NOT EXISTS idx_tasks_assigned_to ON tasks(assigned_to);
    CREATE INDEX IF NOT EXISTS idx_tasks_created_by ON tasks(created_by);

    CREATE TABLE IF NOT EXISTS material_library (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        unique_id TEXT DEFAULT '', product TEXT DEFAULT '',
        material_code TEXT DEFAULT '', material_name TEXT DEFAULT '',
        material_model TEXT DEFAULT '', material_desc TEXT DEFAULT '',
        bom_2k_v2 TEXT DEFAULT '', bom_2k_v5 TEXT DEFAULT '',
        bom_s21_b0_ti TEXT DEFAULT '', bom_s20_b0_rtw TEXT DEFAULT '',
        bom_f_project TEXT DEFAULT '', bom_forward TEXT DEFAULT '',
        bom_general TEXT DEFAULT '', bom_24t24r_sat TEXT DEFAULT '',
        bom_24t24r_soc TEXT DEFAULT '', bom_sfr2_4d TEXT DEFAULT '',
        bom_nio_8t8r TEXT DEFAULT '', bom_horizon_8t8r TEXT DEFAULT '',
        bom_total TEXT DEFAULT '',
        unit TEXT DEFAULT '', manufacturer TEXT DEFAULT '',
        supplier TEXT DEFAULT '', currency TEXT DEFAULT '',
        unit_price_untaxed TEXT DEFAULT '', unit_price_taxed TEXT DEFAULT '',
        spq TEXT DEFAULT '', moq TEXT DEFAULT '', lead_time TEXT DEFAULT '',
        production_time TEXT DEFAULT '', factory_monthly_output TEXT DEFAULT '',
        distributor_monthly_output TEXT DEFAULT '',
        automotive_applications TEXT DEFAULT '', wafer_origin TEXT DEFAULT '',
        packaging_test_origin TEXT DEFAULT '', shipping_origin TEXT DEFAULT '',
        peak_monthly_capacity TEXT DEFAULT '', cancel_window_days TEXT DEFAULT '',
        update_frequency TEXT DEFAULT '', attachment_note TEXT DEFAULT '',
        source_task_id INTEGER REFERENCES tasks(id),
        source_type TEXT DEFAULT 'task' CHECK(source_type IN ('task','manual','import')),
        custom_data TEXT DEFAULT '{{}}',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS custom_fields (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE INDEX IF NOT EXISTS idx_ml_material_code ON material_library(material_code);
    CREATE INDEX IF NOT EXISTS idx_ml_material_name ON material_library(material_name);
    CREATE INDEX IF NOT EXISTS idx_ml_manufacturer ON material_library(manufacturer);
    CREATE INDEX IF NOT EXISTS idx_ml_supplier ON material_library(supplier);
    CREATE INDEX IF NOT EXISTS idx_ml_material_model ON material_library(material_model);
    CREATE INDEX IF NOT EXISTS idx_ml_product ON material_library(product);
    CREATE INDEX IF NOT EXISTS idx_ml_updated_at ON material_library(updated_at);
    CREATE INDEX IF NOT EXISTS idx_ml_source_type ON material_library(source_type);
    """)

    # Migrate existing databases: add security columns if missing
    for col, col_def in [
        ('password_changed', 'INTEGER DEFAULT 0'),
        ('failed_attempts', 'INTEGER DEFAULT 0'),
        ('locked_until', "TEXT DEFAULT ''"),
    ]:
        try:
            db.execute(f"ALTER TABLE users ADD COLUMN {col} {col_def}")
        except sqlite3.OperationalError:
            pass  # Column already exists

    # Migrate: add attachment_file to tasks
    try:
        db.execute("ALTER TABLE tasks ADD COLUMN attachment_file TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    # Migrate: add permissions to users
    try:
        db.execute("ALTER TABLE users ADD COLUMN permissions TEXT DEFAULT '{}'")
    except sqlite3.OperationalError:
        pass

    # Migrate: add custom_data to material_library
    try:
        db.execute("ALTER TABLE material_library ADD COLUMN custom_data TEXT DEFAULT '{}'")
    except sqlite3.OperationalError:
        pass

    count = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if count == 0:
        users = [
            ('leader',   generate_password_hash('leader123'), '王经理', 'leader'),
            ('zhangsan', generate_password_hash('pass123'),   '张三',   'member'),
            ('lisi',     generate_password_hash('pass123'),   '李四',   'member'),
            ('wangwu',   generate_password_hash('pass123'),   '王五',   'member'),
        ]
        db.executemany("INSERT INTO users (username, password_hash, display_name, role) VALUES (?,?,?,?)", users)
        print("[OK] 已创建预置用户（首次登录需修改密码）")

    db.commit()
    db.close()
    print("[OK] 数据库就绪")

# ============================================================
# Security Helpers
# ============================================================
def validate_password_strength(password):
    """返回 (is_valid, error_message)"""
    if len(password) < 8:
        return False, '密码至少需要 8 位'
    has_letter = any(c.isalpha() for c in password)
    has_digit = any(c.isdigit() for c in password)
    if not has_letter:
        return False, '密码必须包含至少一个字母'
    if not has_digit:
        return False, '密码必须包含至少一个数字'
    return True, ''

MAX_FAILED_ATTEMPTS = 5
LOCKOUT_MINUTES = 15

# ============================================================
# Auth
# ============================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def leader_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        if session.get('role') != 'leader':
            flash('仅 Leader 可执行此操作', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def get_current_user():
    if 'user_id' not in session:
        return None
    db = get_db()
    return db.execute("SELECT * FROM users WHERE id=?", [session['user_id']]).fetchone()

def user_can(user, permission):
    """检查用户是否有某项权限（Leader拥有所有权限）"""
    if not user:
        return False
    if user['role'] == 'leader':
        return True
    try:
        perms = json.loads(user['permissions'] or '{}')
    except (json.JSONDecodeError, TypeError):
        perms = {}
    return perms.get(permission, False)

def can_manage_library(user):
    """检查用户是否可以管理物料库（删除、导入等）"""
    return user_can(user, 'manage_library')

@app.context_processor
def inject_globals():
    return dict(current_date=datetime.now().strftime('%Y-%m-%d'))

# ============================================================
# Dashboard Data Builder (for task_dashboard.html)
# ============================================================
def build_dashboard_data(task):
    """从 task row 构建仪表盘所需数据结构"""
    # BOM 数据（柱状图 + 环形图）
    bom_fields = [f for f in MODULES[1]['fields'] if f['col'] != 'bom_total']
    bom_data = []
    for f in bom_fields:
        val = task[f['col']]
        try:
            num_val = float(val) if val else 0
        except (ValueError, TypeError):
            num_val = 0
        bom_data.append({'label': f['label'], 'value': num_val})
    bom_data.sort(key=lambda x: x['value'], reverse=True)

    # 完整度统计
    all_fields = []
    filled = 0
    empty = 0
    for mod in MODULES:
        for f in mod['fields']:
            val = task[f['col']]
            is_filled = bool(val and str(val).strip())
            all_fields.append({
                'category': mod['name'],
                'field': f['label'],
                'value': str(val).strip() if is_filled else None,
                'filled': is_filled
            })
            if is_filled:
                filled += 1
            else:
                empty += 1

    # 供应链数据
    supply_keys = [
        ('manufacturer', '品牌/制造商'), ('supplier', '供应商'),
        ('currency', '币别'), ('unit', '单位'),
        ('unit_price_untaxed', '单价未税'), ('unit_price_taxed', '单价含税'),
        ('spq', 'SPQ（标准包装）'), ('moq', 'MOQ（最小起订）'),
        ('lead_time', 'L/T（交期）'), ('production_time', '量产时间'),
        ('factory_monthly_output', '原厂月出货量'),
        ('distributor_monthly_output', '代理商月出货量'),
        ('automotive_applications', '车厂应用实绩'),
        ('wafer_origin', '晶圆产地'), ('packaging_test_origin', '封测产地'),
        ('shipping_origin', '出货地'), ('peak_monthly_capacity', '峰值月产能'),
        ('cancel_window_days', 'Cancel Window Days'),
    ]
    supply_data = []
    for col, label in supply_keys:
        val = task[col]
        if val and str(val).strip():
            supply_data.append({'label': label, 'value': str(val).strip()})

    total_fields = filled + empty
    pct = round(filled / total_fields * 100, 1) if total_fields > 0 else 0

    return {
        'bom_data': bom_data,
        'all_fields': all_fields,
        'filled': filled,
        'empty': empty,
        'total': total_fields,
        'pct': pct,
        'supply_data': supply_data,
    }

# ============================================================
# Auth Routes
# ============================================================
@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username=?", [username]).fetchone()

        # 检查账号是否被锁定
        if user:
            locked_until = user['locked_until'] or ''
            if locked_until:
                try:
                    lock_time = datetime.fromisoformat(locked_until)
                    if datetime.now() < lock_time:
                        remaining = int((lock_time - datetime.now()).total_seconds() / 60) + 1
                        flash(f'账号已锁定，请 {remaining} 分钟后再试', 'error')
                        return render_template('login.html')
                except (ValueError, TypeError):
                    pass  # Invalid timestamp, ignore

        if user and check_password_hash(user['password_hash'], password):
            # 登录成功：清除锁定记录
            db.execute("UPDATE users SET failed_attempts=0, locked_until='' WHERE id=?",
                       [user['id']])
            db.commit()

            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['display_name'] = user['display_name']
            session.permanent = True

            flash(f'欢迎回来，{user["display_name"]}！', 'success')
            return redirect(url_for('dashboard'))

        # 登录失败：记录失败次数
        if user:
            new_attempts = (user['failed_attempts'] or 0) + 1
            if new_attempts >= MAX_FAILED_ATTEMPTS:
                lock_time = datetime.now() + timedelta(minutes=LOCKOUT_MINUTES)
                db.execute("UPDATE users SET failed_attempts=?, locked_until=? WHERE id=?",
                           [new_attempts, lock_time.isoformat(), user['id']])
                db.commit()
                flash(f'密码连续错误 {MAX_FAILED_ATTEMPTS} 次，账号已锁定 {LOCKOUT_MINUTES} 分钟', 'error')
            else:
                db.execute("UPDATE users SET failed_attempts=? WHERE id=?",
                           [new_attempts, user['id']])
                db.commit()
                remaining = MAX_FAILED_ATTEMPTS - new_attempts
                flash(f'用户名或密码错误，还剩 {remaining} 次尝试机会', 'error')
        else:
            flash('用户名或密码错误', 'error')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

# ============================================================
# Dashboard
# ============================================================
@app.route('/')
@login_required
def dashboard():
    user = get_current_user()
    db = get_db()

    if user['role'] == 'leader':
        total_tasks = db.execute("SELECT COUNT(*) FROM tasks").fetchone()[0]
        my_created = db.execute("SELECT COUNT(*) FROM tasks WHERE created_by=?", [user['id']]).fetchone()[0]
        pending = db.execute("SELECT COUNT(*) FROM tasks WHERE status IN ('assigned','in_progress')").fetchone()[0]
        submitted = db.execute("SELECT COUNT(*) FROM tasks WHERE status='submitted'").fetchone()[0]
        completed = db.execute("SELECT COUNT(*) FROM tasks WHERE status='completed'").fetchone()[0]
        tasks = db.execute("""
            SELECT t.*, u.display_name as assignee_name
            FROM tasks t LEFT JOIN users u ON t.assigned_to = u.id
            ORDER BY CASE t.status WHEN 'submitted' THEN 1 WHEN 'in_progress' THEN 2 WHEN 'assigned' THEN 3 ELSE 4 END, t.updated_at DESC
        """).fetchall()
    else:
        total_tasks = db.execute("SELECT COUNT(*) FROM tasks WHERE assigned_to=?", [user['id']]).fetchone()[0]
        my_created = total_tasks  # 成员视角下"我的任务"=全部
        pending = db.execute("SELECT COUNT(*) FROM tasks WHERE assigned_to=? AND status IN ('assigned','in_progress')", [user['id']]).fetchone()[0]
        submitted = db.execute("SELECT COUNT(*) FROM tasks WHERE assigned_to=? AND status='submitted'", [user['id']]).fetchone()[0]
        completed = db.execute("SELECT COUNT(*) FROM tasks WHERE assigned_to=? AND status='completed'", [user['id']]).fetchone()[0]
        tasks = db.execute("""
            SELECT t.*, u.display_name as assignee_name
            FROM tasks t LEFT JOIN users u ON t.assigned_to = u.id
            WHERE t.assigned_to=?
            ORDER BY CASE t.status WHEN 'submitted' THEN 1 WHEN 'in_progress' THEN 2 WHEN 'assigned' THEN 3 ELSE 4 END, t.updated_at DESC
        """, [user['id']]).fetchall()

    return render_template('dashboard.html',
        user=user, tasks=tasks, STATUS_MAP=STATUS_MAP,
        total_tasks=total_tasks, my_created=my_created,
        pending=pending, submitted=submitted, completed=completed)

# ============================================================
# Create Task (Leader) — 先填物料信息，再指派
# ============================================================
@app.route('/task/new', methods=['GET', 'POST'])
@login_required
@leader_required
def create_task():
    user = get_current_user()
    db = get_db()

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        assigned_to = request.form.get('assigned_to', '')
        due_date = request.form.get('due_date', '')

        if not title:
            flash('请输入任务标题', 'error')
            return redirect(url_for('create_task'))

        # 收集料件信息（LEADER_MODULES[0] 的字段）
        comp_cols = []
        comp_vals = []
        for f in LEADER_MODULES[0]['fields']:
            val = request.form.get(f['col'], '').strip()
            comp_cols.append(f['col'])
            comp_vals.append(val)
            if f.get('required') and not val:
                flash(f'请填写必填字段: {f["label"]}', 'error')
                return redirect(url_for('create_task'))

        # 构建 INSERT：title + 料件信息字段 + assigned_to + created_by + status + due_date
        cols = ['title'] + comp_cols + ['assigned_to', 'created_by', 'status', 'due_date']
        vals = [title] + comp_vals + [assigned_to, user['id'], 'assigned', due_date or None]

        db.execute(
            f"INSERT INTO tasks ({','.join(cols)}) VALUES ({','.join(['?' for _ in vals])})",
            vals
        )
        task_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        assignee_name = db.execute("SELECT display_name FROM users WHERE id=?", [assigned_to]).fetchone()['display_name']
        db.execute("INSERT INTO activity_log (task_id, user_id, action, comment) VALUES (?,?,'created',?)",
                   [task_id, user['id'], f'创建任务，填写料件信息后指派给 {assignee_name}'])
        db.commit()
        flash(f'任务已创建，物料信息已预填，已通知 {assignee_name}！', 'success')
        return redirect(url_for('dashboard'))

    members = db.execute("SELECT * FROM users WHERE role='member' ORDER BY display_name").fetchall()
    return render_template('create_task.html', user=user, members=members,
                          component_fields=LEADER_MODULES[0]['fields'])

# ============================================================
# View Task
# ============================================================
@app.route('/task/<int:task_id>')
@login_required
def view_task(task_id):
    user = get_current_user()
    db = get_db()

    task = db.execute("""
        SELECT t.*, u1.display_name as assignee_name, u2.display_name as creator_name
        FROM tasks t LEFT JOIN users u1 ON t.assigned_to = u1.id
        LEFT JOIN users u2 ON t.created_by = u2.id WHERE t.id=?
    """, [task_id]).fetchone()

    if not task:
        flash('任务不存在', 'error')
        return redirect(url_for('dashboard'))

    if user['role'] != 'leader' and task['assigned_to'] != user['id']:
        flash('无权查看此任务', 'error')
        return redirect(url_for('dashboard'))

    logs = db.execute("""
        SELECT a.*, u.display_name FROM activity_log a
        JOIN users u ON a.user_id = u.id
        WHERE a.task_id=? ORDER BY a.created_at DESC
    """, [task_id]).fetchall()

    return render_template('view_task.html',
        user=user, task=task, logs=logs,
        modules=MODULES, STATUS_MAP=STATUS_MAP)

# ============================================================
# Task Dashboard (standalone route)
# ============================================================
@app.route('/task/<int:task_id>/dashboard')
@login_required
def task_dashboard(task_id):
    """独立仪表盘视图——任何状态的任务都可以查看"""
    user = get_current_user()
    db = get_db()

    task = db.execute("""
        SELECT t.*, u1.display_name as assignee_name, u2.display_name as creator_name
        FROM tasks t LEFT JOIN users u1 ON t.assigned_to = u1.id
        LEFT JOIN users u2 ON t.created_by = u2.id WHERE t.id=?
    """, [task_id]).fetchone()

    if not task:
        flash('任务不存在', 'error')
        return redirect(url_for('dashboard'))

    if user['role'] != 'leader' and task['assigned_to'] != user['id']:
        flash('无权查看此任务', 'error')
        return redirect(url_for('dashboard'))

    dashboard_data = build_dashboard_data(task)
    logs = db.execute("""
        SELECT a.*, u.display_name FROM activity_log a
        JOIN users u ON a.user_id = u.id
        WHERE a.task_id=? ORDER BY a.created_at DESC
    """, [task_id]).fetchall()

    return render_template('task_dashboard.html',
        user=user, task=task, logs=logs,
        modules=MODULES, STATUS_MAP=STATUS_MAP,
        dashboard_data=dashboard_data)

# ============================================================
# Fill Task (Member)
# ============================================================
@app.route('/task/<int:task_id>/fill', methods=['GET', 'POST'])
@login_required
def fill_task(task_id):
    user = get_current_user()
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id=?", [task_id]).fetchone()

    if not task:
        flash('任务不存在', 'error')
        return redirect(url_for('dashboard'))

    # Leader can preview the fill form (read-only)
    is_preview = (user['role'] == 'leader' and task['assigned_to'] != user['id'])

    if not is_preview and task['assigned_to'] != user['id']:
        flash('此任务未分配给你', 'error')
        return redirect(url_for('dashboard'))

    if is_preview and request.method == 'POST':
        flash('预览模式下不能提交数据', 'error')
        return redirect(url_for('view_task', task_id=task_id))

    if request.method == 'POST':
        action = request.form.get('action', 'save')

        updates = {}
        # 只收集 MEMBER_MODULES 的字段（料件信息由 Leader 创建时已填）
        for mod in MEMBER_MODULES:
            for f in mod['fields']:
                col = f['col']
                val = request.form.get(col, '').strip()
                if val:
                    updates[col] = val

        # Auto-calculate BOM total
        bom_cols = [f['col'] for f in MEMBER_MODULES[0]['fields'] if f['col'] != 'bom_total']
        try:
            total = sum(float(updates.get(c, 0) or 0) for c in bom_cols)
            updates['bom_total'] = str(int(total)) if total == int(total) else str(total)
        except:
            pass

        # Handle file upload
        uploaded_file = request.files.get('attachment')
        if uploaded_file and uploaded_file.filename:
            # 保留原扩展名，生成唯一文件名
            import uuid
            ext = os.path.splitext(uploaded_file.filename)[1]
            safe_name = f"{uuid.uuid4().hex}{ext}"
            file_path = os.path.join(UPLOAD_DIR, safe_name)
            uploaded_file.save(file_path)
            # 存储 原始文件名|存储文件名|备注
            note = request.form.get('attachment_note', '').strip()
            updates['attachment_file'] = f"{uploaded_file.filename}|{safe_name}|{note}"

        new_status = 'submitted' if action == 'submit' else 'in_progress'

        set_parts = [f"{col}=?" for col in updates.keys()]
        values = list(updates.values())
        set_parts.append("status=?")
        values.append(new_status)
        set_parts.append("updated_at=CURRENT_TIMESTAMP")

        sql = f"UPDATE tasks SET {', '.join(set_parts)} WHERE id=?"
        values.append(task_id)
        db.execute(sql, values)

        log_action = 'submitted' if action == 'submit' else 'updated'
        log_comment = '提交任务，等待 Leader 审批' if action == 'submit' else '保存草稿'
        db.execute("INSERT INTO activity_log (task_id, user_id, action, comment) VALUES (?,?,?,?)",
                   [task_id, user['id'], log_action, log_comment])
        db.commit()

        flash('任务已提交！等待 Leader 审批。' if action == 'submit' else '草稿已保存。', 'success' if action == 'submit' else 'info')
        return redirect(url_for('view_task', task_id=task_id))

    # Mark in_progress on first entry (only for the actual assignee)
    if task['status'] == 'assigned' and not is_preview:
        db.execute("UPDATE tasks SET status='in_progress', updated_at=CURRENT_TIMESTAMP WHERE id=?", [task_id])
        db.execute("INSERT INTO activity_log (task_id, user_id, action, comment) VALUES (?,?,'started','开始填写')",
                   [task_id, user['id']])
        db.commit()

    return render_template('fill_task.html', user=user, task=task,
                          leader_modules=LEADER_MODULES, member_modules=MEMBER_MODULES, is_preview=is_preview)

# ============================================================
# Complete / Return (Leader)
# ============================================================
@app.route('/task/<int:task_id>/complete', methods=['POST'])
@login_required
@leader_required
def complete_task(task_id):
    user = get_current_user()
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id=?", [task_id]).fetchone()
    if not task or task['status'] != 'submitted':
        flash('只能对"已提交"的任务操作', 'error')
        return redirect(url_for('dashboard'))

    db.execute("UPDATE tasks SET status='completed', updated_at=CURRENT_TIMESTAMP WHERE id=?", [task_id])
    db.execute("INSERT INTO activity_log (task_id, user_id, action, comment) VALUES (?,?,'completed','审批通过 ✓')",
               [task_id, user['id']])

    # 同步到统一物料库（按物料编码去重，存在则更新）
    all_cols = [f['col'] for f in [f for mod in MODULES for f in mod['fields']]]
    vals = {col: (task[col] or '') for col in all_cols}

    existing = db.execute("SELECT id FROM material_library WHERE material_code=? AND material_code!=''",
                          [vals['material_code']]).fetchone()
    if existing:
        set_parts = [f"{col}=?" for col in all_cols]
        set_parts.append("source_task_id=?")
        set_parts.append("source_type=?")
        set_parts.append("updated_at=CURRENT_TIMESTAMP")
        db.execute(f"UPDATE material_library SET {', '.join(set_parts)} WHERE id=?",
                   [vals[c] for c in all_cols] + [task_id, 'task', existing['id']])
    else:
        cols_sql = ', '.join(all_cols + ['source_task_id', 'source_type'])
        placeholders = ', '.join(['?' for _ in range(len(all_cols) + 2)])
        db.execute(f"INSERT INTO material_library ({cols_sql}) VALUES ({placeholders})",
                   [vals[c] for c in all_cols] + [task_id, 'task'])

    db.commit()
    flash('已审批通过！', 'success')
    return redirect(url_for('view_task', task_id=task_id))

@app.route('/task/<int:task_id>/attachment')
@login_required
def download_attachment(task_id):
    """下载任务附件"""
    db = get_db()
    task = db.execute("SELECT attachment_file FROM tasks WHERE id=?", [task_id]).fetchone()
    if not task or not task['attachment_file']:
        flash('该任务没有附件', 'error')
        return redirect(url_for('view_task', task_id=task_id))

    parts = task['attachment_file'].split('|', 1)
    if len(parts) != 2:
        flash('附件记录异常', 'error')
        return redirect(url_for('view_task', task_id=task_id))

    original_name, stored_name = parts
    file_path = os.path.join(UPLOAD_DIR, stored_name)
    if not os.path.exists(file_path):
        flash('附件文件不存在，可能已被删除', 'error')
        return redirect(url_for('view_task', task_id=task_id))

    return send_file(file_path, as_attachment=True, download_name=original_name)

@app.route('/task/<int:task_id>/delete-attachment', methods=['POST'])
@login_required
def delete_attachment(task_id):
    """删除任务附件"""
    user = get_current_user()
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id=?", [task_id]).fetchone()
    if not task:
        flash('任务不存在', 'error')
        return redirect(url_for('dashboard'))

    # 权限：成员只能删自己任务的附件，leader 可以删任何
    if user['role'] != 'leader' and task['assigned_to'] != user['id']:
        flash('无权删除此附件', 'error')
        return redirect(url_for('dashboard'))

    if not task['attachment_file']:
        flash('该任务没有附件', 'error')
        return redirect(url_for('view_task', task_id=task_id))

    # 删除物理文件
    parts = task['attachment_file'].split('|')
    if len(parts) >= 2:
        stored_name = parts[1]
        file_path = os.path.join(UPLOAD_DIR, stored_name)
        try:
            os.remove(file_path)
        except Exception:
            pass

    db.execute("UPDATE tasks SET attachment_file='', updated_at=CURRENT_TIMESTAMP WHERE id=?", [task_id])
    db.execute("INSERT INTO activity_log (task_id, user_id, action, comment) VALUES (?,?,'delete_attachment','🗑 删除了附件')",
               [task_id, user['id']])
    db.commit()
    flash('附件已删除', 'info')
    return redirect(url_for('view_task', task_id=task_id))

@app.route('/task/<int:task_id>/undo-complete', methods=['POST'])
@login_required
@leader_required
def undo_complete_task(task_id):
    """撤销审核：completed → submitted"""
    user = get_current_user()
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id=?", [task_id]).fetchone()
    if not task or task['status'] != 'completed':
        flash('只能撤销"已完成"状态的任务', 'error')
        return redirect(url_for('dashboard'))

    db.execute("UPDATE tasks SET status='submitted', updated_at=CURRENT_TIMESTAMP WHERE id=?", [task_id])
    db.execute("INSERT INTO activity_log (task_id, user_id, action, comment) VALUES (?,?,'undo_complete','↩️ 撤销审核，恢复为已提交状态')",
               [task_id, user['id']])
    db.commit()
    flash('审核已撤销，任务恢复为"已提交"状态。', 'info')
    return redirect(url_for('view_task', task_id=task_id))

@app.route('/task/<int:task_id>/withdraw', methods=['POST'])
@login_required
def withdraw_task(task_id):
    """成员撤回已提交的任务（submitted → in_progress）"""
    user = get_current_user()
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id=?", [task_id]).fetchone()

    if not task:
        flash('任务不存在', 'error')
        return redirect(url_for('dashboard'))
    if task['assigned_to'] != user['id']:
        flash('只能撤回自己的任务', 'error')
        return redirect(url_for('dashboard'))
    if task['status'] != 'submitted':
        flash('只能撤回"已提交"状态的任务', 'error')
        return redirect(url_for('dashboard'))

    db.execute("UPDATE tasks SET status='in_progress', updated_at=CURRENT_TIMESTAMP WHERE id=?", [task_id])
    db.execute("INSERT INTO activity_log (task_id, user_id, action, comment) VALUES (?,?,'withdrawn','撤回提交，重新编辑')",
               [task_id, user['id']])
    db.commit()
    flash('任务已撤回，可以重新编辑和提交。', 'info')
    return redirect(url_for('view_task', task_id=task_id))

@app.route('/task/<int:task_id>/edit', methods=['GET', 'POST'])
@login_required
@leader_required
def edit_task(task_id):
    """Leader编辑任务标题、负责人、截止日期"""
    user = get_current_user()
    db = get_db()
    task = db.execute("SELECT t.*, u.display_name as assignee_name FROM tasks t LEFT JOIN users u ON t.assigned_to=u.id WHERE t.id=?", [task_id]).fetchone()

    if not task:
        flash('任务不存在', 'error')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        assigned_to = request.form.get('assigned_to', '')
        due_date = request.form.get('due_date', '')

        if not title:
            flash('任务标题不能为空', 'error')
            return redirect(url_for('edit_task', task_id=task_id))

        old_assignee = task['assigned_to']
        db.execute("UPDATE tasks SET title=?, assigned_to=?, due_date=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                   [title, assigned_to, due_date or None, task_id])

        changes = []
        if title != task['title']:
            changes.append(f'标题改为: {title}')
        if int(assigned_to) != old_assignee:
            new_name = db.execute("SELECT display_name FROM users WHERE id=?", [assigned_to]).fetchone()['display_name']
            changes.append(f'负责人改为: {new_name}')
        if due_date != (task['due_date'] or ''):
            changes.append(f'截止日期改为: {due_date}')

        db.execute("INSERT INTO activity_log (task_id, user_id, action, comment) VALUES (?,?,'edited',?)",
                   [task_id, user['id'], '; '.join(changes) if changes else '任务信息已更新'])
        db.commit()
        flash('任务信息已更新！', 'success')
        return redirect(url_for('view_task', task_id=task_id))

    members = db.execute("SELECT * FROM users WHERE role='member' ORDER BY display_name").fetchall()
    return render_template('edit_task.html', user=user, task=task, members=members)

@app.route('/task/<int:task_id>/return', methods=['POST'])
@login_required
@leader_required
def return_task(task_id):
    user = get_current_user()
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id=?", [task_id]).fetchone()
    if not task or task['status'] != 'submitted':
        flash('只能对"已提交"的任务操作', 'error')
        return redirect(url_for('dashboard'))

    reason = request.form.get('reason', '').strip()
    db.execute("UPDATE tasks SET status='in_progress', updated_at=CURRENT_TIMESTAMP WHERE id=?", [task_id])
    db.execute("INSERT INTO activity_log (task_id, user_id, action, comment) VALUES (?,?,'returned',?)",
               [task_id, user['id'], f'退回修改: {reason}' if reason else '退回修改'])
    db.commit()
    flash('已退回给负责人修改。', 'info')
    return redirect(url_for('view_task', task_id=task_id))

# ============================================================
# Change Password (all users)
# ============================================================
@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    user = get_current_user()
    db = get_db()

    if request.method == 'POST':
        old_pw = request.form.get('old_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')

        if not check_password_hash(user['password_hash'], old_pw):
            flash('当前密码错误', 'error')
        elif new_pw != confirm_pw:
            flash('两次输入的新密码不一致', 'error')
        else:
            valid, err_msg = validate_password_strength(new_pw)
            if not valid:
                flash(err_msg, 'error')
            else:
                db.execute("UPDATE users SET password_hash=?, password_changed=1 WHERE id=?",
                           [generate_password_hash(new_pw), user['id']])
                db.commit()

                # 清除强制改密标记
                session.pop('force_pw_change', None)
                flash('密码修改成功！', 'success')
                return redirect(url_for('dashboard'))

    force_change = session.get('force_pw_change', False)
    return render_template('change_password.html', user=user, force_change=force_change)

# ============================================================
# User Management (Leader only)
# ============================================================
@app.route('/users')
@login_required
@leader_required
def user_list():
    user = get_current_user()
    db = get_db()
    users = db.execute("SELECT * FROM users ORDER BY role, created_at").fetchall()
    # 为每个用户解析 permissions JSON
    user_list = []
    for u in users:
        u_dict = dict(u)
        try:
            u_dict['perms'] = json.loads(u['permissions'] or '{}')
        except (json.JSONDecodeError, TypeError):
            u_dict['perms'] = {}
        user_list.append(u_dict)
    return render_template('user_list.html', user=user, all_users=user_list)

@app.route('/users/add', methods=['POST'])
@login_required
@leader_required
def user_add():
    user = get_current_user()
    db = get_db()

    username = request.form.get('username', '').strip()
    display_name = request.form.get('display_name', '').strip()
    role = request.form.get('role', 'member')
    password = request.form.get('password', '').strip()

    if not username or not display_name or not password:
        flash('用户名、显示名和密码不能为空', 'error')
    else:
        valid, err_msg = validate_password_strength(password)
        if not valid:
            flash(err_msg, 'error')
        else:
            existing = db.execute("SELECT id FROM users WHERE username=?", [username]).fetchone()
            if existing:
                flash(f'用户名 "{username}" 已存在', 'error')
            else:
                db.execute("INSERT INTO users (username, password_hash, display_name, role) VALUES (?,?,?,?)",
                           [username, generate_password_hash(password), display_name, role])
                db.commit()
                flash(f'用户 {display_name}（{username}）创建成功！', 'success')

    return redirect(url_for('user_list'))

@app.route('/users/<int:uid>/reset-password', methods=['POST'])
@login_required
@leader_required
def user_reset_password(uid):
    db = get_db()
    new_pw = request.form.get('new_password', '').strip()
    target = db.execute("SELECT * FROM users WHERE id=?", [uid]).fetchone()

    if not target:
        flash('用户不存在', 'error')
    else:
        valid, err_msg = validate_password_strength(new_pw)
        if not valid:
            flash(err_msg, 'error')
        else:
            db.execute("UPDATE users SET password_hash=?, password_changed=0 WHERE id=?",
                       [generate_password_hash(new_pw), uid])
            db.commit()
            flash(f'{target["display_name"]} 的密码已重置！下次登录需重新设置密码。', 'success')

    return redirect(url_for('user_list'))

@app.route('/users/<int:uid>/permission', methods=['POST'])
@login_required
@leader_required
def user_permission(uid):
    """Leader 为成员设置权限"""
    db = get_db()
    target = db.execute("SELECT * FROM users WHERE id=?", [uid]).fetchone()
    if not target:
        return jsonify({'error': '用户不存在'}), 404
    if target['role'] == 'leader':
        return jsonify({'error': 'Leader 拥有全部权限，无需单独设置'}), 400

    data = request.get_json()
    perm_name = data.get('permission', '')
    perm_value = data.get('value', False)

    try:
        perms = json.loads(target['permissions'] or '{}')
    except (json.JSONDecodeError, TypeError):
        perms = {}
    perms[perm_name] = bool(perm_value)

    db.execute("UPDATE users SET permissions=? WHERE id=?",
               [json.dumps(perms, ensure_ascii=False), uid])
    db.commit()
    return jsonify({'success': True})

# ============================================================
# API — Stats polling
# ============================================================
@app.route('/api/stats')
@login_required
def api_stats():
    user = get_current_user()
    db = get_db()
    if user['role'] == 'leader':
        n = db.execute("SELECT COUNT(*) FROM tasks WHERE status='submitted'").fetchone()[0]
    else:
        n = db.execute("SELECT COUNT(*) FROM tasks WHERE assigned_to=? AND status IN ('assigned','in_progress')", [user['id']]).fetchone()[0]
    return jsonify({'pending': n})

# ============================================================
# Excel 导出工具函数
# ============================================================
def _build_excel(rows, fields=None):
    """生成 Excel 工作簿，rows 可以是 sqlite Row 或 dict 列表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "电子料台账"

    if fields is None:
        fields = [f for mod in MODULES for f in mod['fields']]

    # Header style
    header_fill = PatternFill(start_color="2A78D6", end_color="2A78D6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name="微软雅黑")
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    cell_font = Font(size=10, name="微软雅黑")
    center_align = Alignment(horizontal='center', vertical='center')

    # Row 1: Module group headers (merged where same module)
    # Row 2: Field labels
    prev_mod = None
    merge_start = 1
    for col_idx, f in enumerate(fields, 1):
        mod_name = None
        for mod in MODULES:
            if f['col'] in [mf['col'] for mf in mod['fields']]:
                mod_name = mod['name']
                break

        cell_hdr = ws.cell(row=1, column=col_idx, value='')
        cell_lbl = ws.cell(row=2, column=col_idx, value=f['label'])
        cell_lbl.fill = header_fill
        cell_lbl.font = header_font
        cell_lbl.border = thin_border
        cell_lbl.alignment = center_align

        if mod_name != prev_mod:
            if prev_mod and merge_start < col_idx - 1:
                ws.merge_cells(start_row=1, start_column=merge_start,
                               end_row=1, end_column=col_idx - 1)
            merge_start = col_idx
            prev_mod = mod_name

    # Row 1 merge last group
    if prev_mod and merge_start <= len(fields):
        ws.merge_cells(start_row=1, start_column=merge_start,
                       end_row=1, end_column=len(fields))

    # Module group headers
    prev_mod = None
    merge_start = 1
    mod_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    mod_font = Font(bold=True, size=11, color="1A56C7", name="微软雅黑")
    for col_idx, f in enumerate(fields, 1):
        mod_name = None
        for mod in MODULES:
            if f['col'] in [mf['col'] for mf in mod['fields']]:
                mod_name = mod['name']
                break
        if mod_name != prev_mod:
            cell = ws.cell(row=1, column=col_idx, value=mod_name)
            cell.fill = mod_fill
            cell.font = mod_font
            cell.border = thin_border
            cell.alignment = center_align
            prev_mod = mod_name

    # Data rows
    for row_idx, row in enumerate(rows, 3):
        for col_idx, f in enumerate(fields, 1):
            val = row[f['col']] if row[f['col']] else ''
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            cell.font = cell_font
            cell.border = thin_border

    # Auto-width
    for col_idx in range(1, len(fields) + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_val in row:
                if cell_val:
                    # CJK characters are ~2x width
                    char_len = sum(2 if ord(c) > 127 else 1 for c in str(cell_val))
                    max_len = max(max_len, char_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # Freeze panes (freeze header rows)
    ws.freeze_panes = 'A3'

    return wb

def _excel_response(wb, filename):
    """将工作簿转为 Flask send_file 响应"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ============================================================
# Excel 导出路由
# ============================================================
@app.route('/export/task/<int:task_id>')
@login_required
def export_task(task_id):
    user = get_current_user()
    db = get_db()
    task = db.execute("SELECT * FROM tasks WHERE id=?", [task_id]).fetchone()
    if not task:
        flash('任务不存在', 'error')
        return redirect(url_for('dashboard'))
    if user['role'] != 'leader' and task['assigned_to'] != user['id']:
        flash('无权访问', 'error')
        return redirect(url_for('dashboard'))

    wb = _build_excel([task])
    filename = f"任务_{task_id}_{task['title']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return _excel_response(wb, filename)

@app.route('/export/tasks')
@login_required
def export_tasks():
    user = get_current_user()
    db = get_db()
    status_filter = request.args.get('status', '')
    mine_only = request.args.get('mine', '0')

    conditions = []
    params = []
    if status_filter:
        conditions.append("t.status=?")
        params.append(status_filter)
    if user['role'] == 'leader':
        if mine_only == '1':
            conditions.append("t.created_by=?")
            params.append(user['id'])
    else:
        conditions.append("t.assigned_to=?")
        params.append(user['id'])

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    tasks = db.execute(f"SELECT t.* FROM tasks t {where} ORDER BY t.updated_at DESC", params).fetchall()

    wb = _build_excel(tasks)
    filename = f"电子料台账_任务导出_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return _excel_response(wb, filename)

@app.route('/export/materials')
@login_required
def export_materials():
    user = get_current_user()
    db = get_db()

    search = request.args.get('q', '').strip()
    if search:
        like = f"%{search}%"
        materials = db.execute("""
            SELECT * FROM material_library
            WHERE material_code LIKE ? OR material_name LIKE ? OR material_model LIKE ?
               OR manufacturer LIKE ? OR supplier LIKE ? OR material_desc LIKE ?
               OR unique_id LIKE ?
            ORDER BY updated_at DESC
        """, [like]*7).fetchall()
    else:
        materials = db.execute("SELECT * FROM material_library ORDER BY updated_at DESC").fetchall()

    wb = _build_excel(materials)
    filename = f"电子料台账_物料库导出_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return _excel_response(wb, filename)

# ============================================================
# Excel 导入路由
# ============================================================
@app.route('/import', methods=['GET'])
@login_required
def import_page():
    """导入向导页面"""
    user = get_current_user()
    return render_template('import_excel.html', user=user, modules=MODULES)

@app.route('/import/upload', methods=['POST'])
@login_required
def import_upload():
    """解析上传的 Excel 文件，自动识别表头结构，返回匹配结果"""
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '请上传 .xlsx 或 .xls 文件'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'无法读取 Excel 文件: {str(e)}'}), 400
    try:

        # ====== 简单可靠的标题行检测 ======
        # 扫描前5行，找包含最多字段关键词的行作为标题行
        all_fields = [f for mod in MODULES for f in mod['fields']]
        FIELD_KW = ['编码', '名称', '型号', '描述', '产品', '品牌', '制造商', '供应商',
                    '单价', '币别', '单位', '产地', '用量', '频次', '交期']

        best_row_idx = 0
        best_kw_count = 0
        for r_idx in range(min(5, ws.max_row)):
            row_text = ' '.join([str(cell.value or '').strip() for cell in ws[r_idx + 1]])
            kw_count = sum(1 for kw in FIELD_KW if kw in row_text)
            if kw_count > best_kw_count:
                best_kw_count = kw_count
                best_row_idx = r_idx

        headers = [str(cell.value or '').strip() for cell in ws[best_row_idx + 1]]
        data_start_row = best_row_idx + 2

        # 清理尾部空列
        while headers and not headers[-1]:
            headers.pop()

        if not headers or best_kw_count == 0:
            return jsonify({'error': '未检测到有效的列标题。请确保Excel包含"物料编码""物料名称"等字段名。'}), 400

        # ====== 自动匹配 ======

        mapping = {}
        all_matched = True  # 是否全部列都高置信度匹配
        for idx, header in enumerate(headers):
            if not header:
                continue
            best_match = None
            best_score = 0
            hdr_clean = header.replace('（', '(').replace('）', ')').replace(' ', '').replace('\n', '').replace('\r', '').lower()

            for f in all_fields:
                lbl_clean = f['label'].replace('（', '(').replace('）', ')').replace(' ', '').replace('\n', '').lower()
                score = 0

                if hdr_clean == lbl_clean:
                    score = 100
                elif hdr_clean in lbl_clean or lbl_clean in hdr_clean:
                    score = 85
                else:
                    hdr_chars = set(hdr_clean)
                    lbl_chars = set(lbl_clean)
                    overlap = hdr_chars & lbl_chars
                    if overlap:
                        union = hdr_chars | lbl_chars
                        ratio = len(overlap) / len(union) if union else 0
                        bonus = 0
                        core_keywords = ['编码', '名称', '型号', '描述', '品牌', '制造商', '供应商',
                                         '单价', '币别', '单位', '产地', '交期', '产能', '用量', '物料']
                        for kw in core_keywords:
                            if kw in hdr_clean and kw in lbl_clean:
                                bonus += 15
                        score = int(ratio * 60) + bonus

                if score > best_score:
                    best_score = score
                    best_match = f

            if best_score >= 30:
                mapping[header] = {'col': best_match['col'], 'label': best_match['label'], 'category': '', 'matched': best_score >= 70}
                if best_score < 70:
                    all_matched = False
            else:
                mapping[header] = {'col': '', 'label': header, 'category': '', 'matched': False}
                all_matched = False

        # 补充 category 信息
        for header, info in mapping.items():
            for mod in MODULES:
                for f in mod['fields']:
                    if f['col'] == info['col']:
                        info['category'] = mod['name']
                        break

        # ====== 读取所有数据行 ======
        all_rows = []
        for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=True):
            row_data = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = str(val).strip() if val is not None else ''
            # 过滤：至少有一个字段有实际内容
            if any(v for v in row_data.values()):
                all_rows.append(row_data)

        # ====== 缓存全部数据 ======
        cache_id = secrets.token_hex(16)
        cache_path = os.path.join(tempfile.gettempdir(), f'ml_import_{cache_id}.json')
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump({'headers': headers, 'rows': all_rows}, f, ensure_ascii=False)
        session['import_cache_id'] = cache_id

        preview = all_rows[:20]
        field_options = [{'col': f['col'], 'label': f['label'], 'category': mod['name']} for mod in MODULES for f in mod['fields']]

        matched_count = sum(1 for v in mapping.values() if v['matched'])
        total_count = len([h for h in headers if h])

        return jsonify({
            'headers': headers,
            'mapping': mapping,
            'preview': preview,
            'total_rows': len(all_rows),
            'cache_id': cache_id,
            'field_options': field_options,
            'modules': [{'key': m['key'], 'name': m['name']} for m in MODULES],
            'auto_import': all_matched,
            'matched_count': matched_count,
            'total_columns': total_count,
            'data_start_row': data_start_row,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"解析失败: {str(e)}"}), 400
@app.route('/import/confirm', methods=['POST'])
@login_required
def import_confirm():
    """确认导入：从缓存文件读取全部数据写入 material_library"""
    user = get_current_user()
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '无效请求'}), 400

    mapping = data.get('mapping', {})  # {header: col_name}
    cache_id = data.get('cache_id', '')

    # 从临时文件读取全部行数据
    cache_path = os.path.join(tempfile.gettempdir(), f'ml_import_{cache_id}.json')
    try:
        with open(cache_path, 'r', encoding='utf-8') as f:
            cached = json.load(f)
        all_rows = cached.get('rows', [])
    except Exception:
        return jsonify({'error': '缓存数据已过期，请重新上传文件'}), 400
    finally:
        # 清理临时文件
        try:
            os.remove(cache_path)
        except Exception:
            pass

    if not all_rows:
        return jsonify({'error': '没有要导入的数据'}), 400

    all_cols = [f['col'] for f in [f for mod in MODULES for f in mod['fields']]]

    imported = 0
    for row_data in all_rows:
        # 根据映射构建物料数据
        vals = {col: '' for col in all_cols}
        for header, col_name in mapping.items():
            if col_name and col_name in vals:
                vals[col_name] = str(row_data.get(header, '')).strip()

        cols_sql = ', '.join(all_cols + ['source_type'])
        placeholders = ', '.join(['?' for _ in range(len(all_cols) + 1)])
        db.execute(f"INSERT INTO material_library ({cols_sql}) VALUES ({placeholders})",
                   [vals[c] for c in all_cols] + ['import'])
        imported += 1

    db.commit()
    return jsonify({
        'success': True,
        'imported': imported,
        'message': f'导入完成：新增 {imported} 条'
    })

# ============================================================
# 统一数据中心 —— 物料库
# ============================================================
@app.route('/materials')
@login_required
def material_library():
    user = get_current_user()
    db = get_db()

    search = request.args.get('q', '').strip()
    product = request.args.get('product', '').strip()
    manufacturer = request.args.get('manufacturer', '').strip()
    supplier = request.args.get('supplier', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 50

    conditions = []
    params = []

    if search:
        like = f"%{search}%"
        cond = ("(material_code LIKE ? OR material_name LIKE ? OR material_model LIKE ?"
                " OR manufacturer LIKE ? OR supplier LIKE ? OR material_desc LIKE ?"
                " OR unique_id LIKE ?)")
        conditions.append(cond)
        params.extend([like] * 7)
    if product:
        conditions.append("product LIKE ?")
        params.append(f"%{product}%")
    if manufacturer:
        conditions.append("manufacturer LIKE ?")
        params.append(f"%{manufacturer}%")
    if supplier:
        conditions.append("supplier LIKE ?")
        params.append(f"%{supplier}%")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    total = db.execute(f"SELECT COUNT(*) FROM material_library {where}", params).fetchone()[0]
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    offset = (page - 1) * per_page

    materials = db.execute(
        f"SELECT * FROM material_library {where} ORDER BY updated_at DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()

    SOURCE_LABELS = {'task': '任务同步', 'manual': '手动录入', 'import': 'Excel导入'}

    # 转为 dict 列表供前端 JS 使用
    materials_json = json.dumps([dict(m) for m in materials], ensure_ascii=False)

    return render_template('material_library.html',
        user=user, materials=materials, modules=MODULES,
        total=total, page=page, total_pages=total_pages, per_page=per_page,
        search=search, product=product, manufacturer=manufacturer, supplier=supplier,
        SOURCE_LABELS=SOURCE_LABELS,
        materials_json=materials_json,
        can_manage=can_manage_library(user))

@app.route('/materials/<int:mat_id>')
@login_required
def material_detail(mat_id):
    """物料详情（JSON）"""
    user = get_current_user()
    db = get_db()
    mat = db.execute("SELECT * FROM material_library WHERE id=?", [mat_id]).fetchone()
    if not mat:
        return jsonify({'error': '物料不存在'}), 404

    fields_data = []
    for mod in MODULES:
        fdata = {'name': mod['name'], 'icon': mod['icon'], 'fields': []}
        for f in mod['fields']:
            fdata['fields'].append({'label': f['label'], 'value': mat[f['col']] or ''})
        fields_data.append(fdata)

    return jsonify({
        'id': mat['id'],
        'fields': fields_data,
        'source_type': mat['source_type'],
        'source_task_id': mat['source_task_id'],
        'created_at': mat['created_at'],
        'updated_at': mat['updated_at'],
        'custom_data': json.loads(mat['custom_data'] or '{}'),
        'custom_fields': [{'id': f['id'], 'name': f['name']} for f in
                          db.execute("SELECT * FROM custom_fields ORDER BY created_at").fetchall()],
    })

@app.route('/materials/<int:mat_id>/dashboard')
@login_required
def material_dashboard(mat_id):
    """物料分析仪表盘——不依赖任务，直接从物料库数据生成"""
    user = get_current_user()
    db = get_db()
    mat = db.execute("SELECT * FROM material_library WHERE id=?", [mat_id]).fetchone()
    if not mat:
        flash('物料不存在', 'error')
        return redirect(url_for('material_library'))

    # build_dashboard_data 需要的字段名和 material_library/tasks 一致，直接复用
    dashboard_data = build_dashboard_data(mat)
    SOURCE_LABELS = {'task': '任务同步', 'manual': '手动录入', 'import': 'Excel导入'}

    return render_template('task_dashboard.html',
        user=user,
        task=mat,  # 复用 task_dashboard 模板，material_library 字段名和 tasks 一样
        modules=MODULES,
        STATUS_MAP={'completed': {'label': '物料库记录', 'color': '#0ca30c', 'bg': '#ecfdf3'}},
        dashboard_data=dashboard_data,
        source_label=SOURCE_LABELS.get(mat['source_type'] or 'task', mat['source_type']),
        is_material=True,  # 标记这是物料库仪表盘
        logs=[])

@app.route('/materials/<int:mat_id>/delete', methods=['POST'])
@login_required
def material_delete(mat_id):
    user = get_current_user()
    if not can_manage_library(user):
        flash('无权限：需要物料库管理权限', 'error')
        return redirect(url_for('material_library'))
    db = get_db()
    mat = db.execute("SELECT * FROM material_library WHERE id=?", [mat_id]).fetchone()
    if not mat:
        flash('物料不存在', 'error')
    else:
        db.execute("DELETE FROM material_library WHERE id=?", [mat_id])
        db.commit()
        flash(f'物料 "{mat["material_name"] or mat["material_code"] or mat_id}" 已删除', 'success')
    return redirect(url_for('material_library'))

# ============================================================
# 筛选自动补全 API
# ============================================================
@app.route('/api/custom-fields', methods=['GET', 'POST', 'DELETE'])
@login_required
def api_custom_fields():
    """管理自定义字段：GET 返回列表，POST 添加，DELETE 删除"""
    user = get_current_user()
    if not can_manage_library(user):
        return jsonify({'error': '无权限'}), 403
    db = get_db()

    if request.method == 'GET':
        fields = db.execute("SELECT * FROM custom_fields ORDER BY created_at").fetchall()
        return jsonify([{'id': f['id'], 'name': f['name']} for f in fields])

    elif request.method == 'POST':
        data = request.get_json()
        name = (data.get('name') or '').strip()
        if not name or len(name) > 30:
            return jsonify({'error': '字段名不能为空且不超过30字'}), 400
        existing = db.execute("SELECT id FROM custom_fields WHERE name=?", [name]).fetchone()
        if existing:
            return jsonify({'error': f'字段「{name}」已存在'}), 400
        db.execute("INSERT INTO custom_fields (name) VALUES (?)", [name])
        db.commit()
        fid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        return jsonify({'id': fid, 'name': name})

    elif request.method == 'DELETE':
        fid = request.args.get('id', type=int)
        if not fid:
            return jsonify({'error': '缺少字段ID'}), 400
        db.execute("DELETE FROM custom_fields WHERE id=?", [fid])
        db.commit()
        return jsonify({'success': True})

@app.route('/api/materials/<int:mat_id>/custom-data', methods=['POST'])
@login_required
def update_custom_data(mat_id):
    """更新某条物料的自定义数据"""
    user = get_current_user()
    if not can_manage_library(user):
        return jsonify({'error': '无权限'}), 403
    db = get_db()
    mat = db.execute("SELECT id, custom_data FROM material_library WHERE id=?", [mat_id]).fetchone()
    if not mat:
        return jsonify({'error': '物料不存在'}), 404

    data = request.get_json()
    try:
        current = json.loads(mat['custom_data'] or '{}')
    except json.JSONDecodeError:
        current = {}

    field_name = (data.get('field') or '').strip()
    value = (data.get('value') or '').strip()

    if not field_name:
        return jsonify({'error': '缺少字段名'}), 400

    current[field_name] = value
    db.execute("UPDATE material_library SET custom_data=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
               [json.dumps(current, ensure_ascii=False), mat_id])
    db.commit()
    return jsonify({'success': True, 'value': value})

@app.route('/api/filter-options')
@login_required
def api_filter_options():
    """返回筛选字段的自动补全选项（限制返回数量，支持搜索）"""
    field = request.args.get('field', '')
    q = request.args.get('q', '').strip()
    db = get_db()

    allowed = {'product': 'product', 'manufacturer': 'manufacturer', 'supplier': 'supplier'}
    if field not in allowed:
        return jsonify({'options': []})

    col = allowed[field]
    if q:
        results = db.execute(
            f"SELECT DISTINCT {col} FROM material_library WHERE {col} LIKE ? AND {col}!='' ORDER BY {col} LIMIT 30",
            [f'%{q}%'])
    else:
        # 无搜索词时返回最常见的30个
        results = db.execute(
            f"SELECT {col}, COUNT(*) as cnt FROM material_library WHERE {col}!='' GROUP BY {col} ORDER BY cnt DESC LIMIT 30")
    return jsonify({'options': [r[0] for r in results]})

# ============================================================
# 全局搜索
# ============================================================
@app.route('/api/search')
@login_required
def api_search():
    """即时搜索API（返回JSON，用于搜索建议）"""
    q = request.args.get('q', '').strip()
    if len(q) < 1:
        return jsonify({'results': []})

    like = f"%{q}%"
    db = get_db()
    results = db.execute("""
        SELECT id, material_code, material_name, material_model, manufacturer, supplier
        FROM material_library
        WHERE material_code LIKE ? OR material_name LIKE ? OR material_model LIKE ?
           OR manufacturer LIKE ? OR supplier LIKE ? OR material_desc LIKE ?
           OR unique_id LIKE ?
        ORDER BY updated_at DESC LIMIT 20
    """, [like]*7).fetchall()

    return jsonify({'results': [{
        'id': r['id'],
        'material_code': r['material_code'],
        'material_name': r['material_name'],
        'material_model': r['material_model'],
        'manufacturer': r['manufacturer'],
        'supplier': r['supplier'],
    } for r in results]})

# ============================================================
# Main
# ============================================================
if __name__ == '__main__':
    print("=" * 60)
    print("  电子料台账协作管理平台")
    print("=" * 60)
    init_db()
    print()
    print("  启动地址: http://localhost:5000")
    print("  Leader:  leader / leader123")
    print("  成员:    zhangsan / pass123")
    print("  成员:    lisi / pass123")
    print("  成员:    wangwu / pass123")
    print("=" * 60)

    # 默认用 waitress 生产模式；加 --dev 参数则用 Flask 开发模式
    use_dev = '--dev' in sys.argv
    if use_dev:
        print("  [DEV模式] 仅调试用，并发弱，勿用于多人协作")
        print("=" * 60)
        app.run(host='0.0.0.0', port=5000, debug=True)
    else:
        from waitress import serve
        print("  [生产模式] waitress 服务已启动，3-4人协作无压力")
        print("=" * 60)
        serve(app, host='0.0.0.0', port=5000)
